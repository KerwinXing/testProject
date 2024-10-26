# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import openpyxl


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.


def read_excel():
    wb = openpyxl.load_workbook('./data/score.xlsx')
    sheet = wb['Sheet1']
    # max_row = sheet.max_row
    # max_col = sheet.max_column
    # print(max_row, max_col)
    result = dict()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        class_name = row[0]
        student_name = row[1]
        score = row[2]
        if score >= 90:
            if class_name not in result:
                result[class_name] = list()
            result[class_name].append(student_name)

    print(result)
    write_txt(result)


def write_txt(data):
    # 打开文件，以写入模式
    wt = open('./data/result.txt', 'w', encoding='utf-8')
    # 遍历data字典
    for key, value in data.items():
        # # 写入key
        # wt.write(f'{key}:\n')
        wt.write(f'{key}: {", ".join(value)}\n')
        # 遍历value
        # for item in value:
            # 写入key和item
            #wt.write(f'{key}\t{item}\n')
    # 关闭文件
    wt.close()


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    # print_hi('PyCharm')
    read_excel()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
